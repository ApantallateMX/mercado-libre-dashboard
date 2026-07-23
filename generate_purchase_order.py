# -*- coding: utf-8 -*-
"""
generate_purchase_order.py — Orden de compra de los SKUs más vendidos entre
TODAS las cuentas ML + Amazon (Apantallate MX).

Por qué existe: la primera vez que se pidió este reporte (2026-07-15) tomó
mucho tiempo porque hubo que: inspeccionar el schema de tokens.db, descubrir
que /api/metrics/top-products cachea SOLO 90 días, descubrir que la API de
ML corta resultados en 10,000 (hay que paginar por rangos de fecha), y
descubrir que la SP-API de Amazon tiene rate limit muy estricto (burst 20,
refill 0.0167 req/s en /orders/v0/orders) que se agota fácil y tarda ~20min
en regenerar. Este script encapsula todo eso para que la próxima vez sea
cuestión de segundos/minutos, no de media sesión de exploración.

Uso:
    py generate_purchase_order.py                        # YTD (desde 1-ene), top 50
    py generate_purchase_order.py --days 90               # últimos 90 días
    py generate_purchase_order.py --since 2026-03-01       # desde fecha específica
    py generate_purchase_order.py --top 20                # solo top 20 SKUs
    py generate_purchase_order.py --amazon-live            # intenta refrescar Amazon
                                                            # en vivo (puede tardar/fallar
                                                            # por rate limit — ver nota abajo)

Amazon — por qué default es "cache" y no "vivo":
    order_history ya tiene un histórico de Amazon guardado por el job de
    background normal del dashboard (_save_amazon_orders_bg). Refrescarlo
    en vivo requiere 1 request a /orders/v0/orders (rate limit compartido
    con TODO lo demás que usa esa cuenta) + 1 request a /orderItems POR
    ORDEN (rate limit aún más estricto). Para un rango de meses con miles
    de órdenes esto puede tardar mucho o fallar directo con 429
    QuotaExceeded si algo más (gap scans, listing checks del dashboard)
    ya venía usando la cuota. --amazon-live lo intenta de todas formas,
    con un presupuesto de tiempo (--amazon-timeout, default 90s) — si se
    agota o falla, cae automáticamente al cache existente.

Salida: orden_compra_top{N}_{periodo}.csv y .xlsx en la raíz del proyecto.
"""
import argparse
import asyncio
import csv
import json
import sqlite3
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

BASE = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE))

DB_PATH = str(BASE / "tokens.db")
ML_ORDER_SEARCH_CAP = 9500  # margen de seguridad bajo el límite real (~10,000) de ML /orders/search

_EXCL_ML_STATUS = {"cancelled", "payment_required", "payment_in_process"}
_EXCL_AMZ_STATUS = {"Cancelled", "Pending"}


# ─────────────────────────────────────────────────────────────────────────
# ML — fetch adaptativo (parte el rango de fechas si supera el cap de ML)
# ─────────────────────────────────────────────────────────────────────────

async def _fetch_ml_range(client, d_from: date, d_to: date, depth=0) -> list:
    df, dt = d_from.isoformat(), d_to.isoformat()
    first = await client.get_orders(offset=0, limit=1, date_from=df, date_to=dt)
    total = first.get("paging", {}).get("total", 0)
    print(f"{'  ' * depth}[ML {client.user_id}] {df} -> {dt}: {total} órdenes")
    if total <= ML_ORDER_SEARCH_CAP or d_from == d_to:
        return await client.fetch_all_orders(date_from=df, date_to=dt)
    mid = d_from + (d_to - d_from) // 2
    left = await _fetch_ml_range(client, d_from, mid, depth + 1)
    right = await _fetch_ml_range(client, mid + timedelta(days=1), d_to, depth + 1)
    return left + right


def _aggregate_ml_orders(orders: list) -> dict:
    """sku -> {units, revenue, title}"""
    sku_sales = {}
    for order in orders:
        if order.get("status") in _EXCL_ML_STATUS:
            continue
        for oi in order.get("order_items", []):
            item = oi.get("item", {})
            sku = (item.get("seller_sku") or "").strip().upper()
            if not sku:
                continue
            title = (item.get("title") or "")[:80]
            qty = oi.get("quantity", 1)
            revenue = (oi.get("unit_price") or 0) * qty
            row = sku_sales.setdefault(sku, {"units": 0, "revenue": 0.0, "title": ""})
            row["units"] += qty
            row["revenue"] += revenue
            if title and not row["title"]:
                row["title"] = title
    return sku_sales


async def fetch_ml_all_accounts(date_from: date, date_to: date) -> dict:
    """nickname -> {sku -> {units, revenue, title}}"""
    from app.services import token_store
    from app.services.meli_client import MeliClient

    accounts = await token_store.get_all_tokens()
    results = {}
    for acc in accounts:
        uid = acc["user_id"]
        nick = acc.get("nickname") or uid
        tok = await token_store.get_tokens(uid)
        if not tok:
            print(f"[ML] {nick}: sin token, se omite")
            continue
        client = MeliClient(tok["access_token"], tok["refresh_token"], uid)
        try:
            orders = await _fetch_ml_range(client, date_from, date_to)
            dedup = {o.get("id"): o for o in orders}
            orders = list(dedup.values())
            results[nick] = _aggregate_ml_orders(orders)
            print(f"[ML] {nick}: {len(orders)} órdenes, {len(results[nick])} SKUs")
        except Exception as e:
            print(f"[ML] {nick}: ERROR {e}")
        finally:
            await client.close()
    return results


# ─────────────────────────────────────────────────────────────────────────
# Amazon — cache (rápido, default) o vivo con presupuesto de tiempo
# ─────────────────────────────────────────────────────────────────────────

def amazon_from_cache(conn: sqlite3.Connection, date_from: date = None, date_to: date = None) -> tuple[dict, dict]:
    """Lee order_history. Devuelve (nickname -> {sku->{units,revenue,title}}, meta)."""
    cur = conn.cursor()
    where = "platform='amazon'"
    params = []
    if date_from and date_to:
        where += " AND order_date BETWEEN ? AND ?"
        params += [date_from.isoformat(), date_to.isoformat()]
    cur.execute(f"SELECT account_id, sku, quantity, unit_price FROM order_history WHERE {where}", params)
    results = {}
    for account_id, sku, qty, price in cur.fetchall():
        sku = (sku or "").strip().upper()
        if not sku:
            continue
        row = results.setdefault(account_id, {}).setdefault(sku, {"units": 0, "revenue": 0.0, "title": ""})
        row["units"] += qty
        row["revenue"] += (price or 0) * qty
    meta_row = cur.execute(
        "SELECT MIN(order_date), MAX(order_date), COUNT(DISTINCT account_id) FROM order_history WHERE platform='amazon'"
    ).fetchone()
    meta = {"min_date": meta_row[0], "max_date": meta_row[1], "n_accounts": meta_row[2], "source": "cache"}
    return results, meta


async def fetch_amazon_all_accounts_live(date_from: date, date_to: date, timeout_budget: int) -> dict:
    """Intenta vivo con presupuesto de tiempo. nickname -> {sku->{units,revenue,title}} (solo cuentas que sí completaron)."""
    from app.services import token_store
    from app.services.amazon_client import get_amazon_client

    accounts = await token_store.get_all_amazon_accounts()
    results = {}
    start = time.time()
    for acc in accounts:
        if time.time() - start > timeout_budget:
            print(f"[AMZ] presupuesto de tiempo agotado ({timeout_budget}s) — resto usará cache")
            break
        seller_id = acc.get("seller_id")
        nick = acc.get("nickname") or seller_id
        client = await get_amazon_client(seller_id=seller_id)
        if not client:
            continue
        try:
            orders = await client.fetch_orders_range(date_from.isoformat(), date_to.isoformat())
            print(f"[AMZ] {nick}: {len(orders)} órdenes en rango — bajando items...")
            sku_sales = {}
            for order in orders:
                if time.time() - start > timeout_budget:
                    print(f"[AMZ] {nick}: presupuesto agotado a mitad de camino, resultado parcial")
                    break
                if order.get("OrderStatus") in _EXCL_AMZ_STATUS:
                    continue
                order_id = order.get("AmazonOrderId", "")
                if not order_id:
                    continue
                try:
                    items = await client.get_order_items(order_id)
                except Exception as e:
                    print(f"[AMZ] {nick}: error items orden {order_id}: {e}")
                    continue
                for it in items:
                    sku = (it.get("SellerSKU") or "").strip().upper()
                    qty = int(it.get("QuantityOrdered") or 0)
                    if not sku or qty <= 0:
                        continue
                    price = float((it.get("ItemPrice") or {}).get("Amount") or 0)
                    row = sku_sales.setdefault(sku, {"units": 0, "revenue": 0.0, "title": ""})
                    row["units"] += qty
                    row["revenue"] += price
            results[nick] = sku_sales
            print(f"[AMZ] {nick}: {len(sku_sales)} SKUs (vivo)")
        except Exception as e:
            print(f"[AMZ] {nick}: FALLÓ en vivo ({e}) — se usará cache para esta cuenta")
        finally:
            await client.close()
    return results


# ─────────────────────────────────────────────────────────────────────────
# BM stock cache (solo lectura — NUNCA llamar BM en vivo, regla del proyecto)
# ─────────────────────────────────────────────────────────────────────────

def bm_stock(conn: sqlite3.Connection, sku: str):
    cur = conn.cursor()
    cur.execute("SELECT data_json, synced_at FROM bm_stock_cache WHERE sku=?", (sku,))
    r = cur.fetchone()
    if not r or (r[1] or 0) <= 0:
        return None, None
    try:
        d = json.loads(r[0])
    except Exception:
        return None, None
    return d.get("avail_total"), d.get("reserved_total")


def seasonal_boost(conn: sqlite3.Connection, title: str = "") -> tuple:
    """Mismo criterio que token_store.get_active_seasonal_boost() pero vía
    sqlite3 sync — evita reimplementar la tabla, solo el query. Retorna
    (multiplier, event_name); (1.0, None) si no hay evento vigente."""
    from datetime import date, timedelta
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT name, start_date, end_date, lead_days, multiplier, category_filter "
            "FROM seasonal_events WHERE active = 1"
        )
        rows = cur.fetchall()
    except sqlite3.OperationalError:
        return 1.0, None  # tabla no existe aún (DB vieja) — sin boost
    today = date.today()
    best_mult, best_name = 1.0, None
    for name, start_s, end_s, lead_days, mult, cat_filter in rows:
        cat_filter = (cat_filter or "").strip()
        if cat_filter and cat_filter.upper() not in (title or "").upper():
            continue
        try:
            start = date.fromisoformat(start_s)
            end = date.fromisoformat(end_s)
        except (ValueError, TypeError):
            continue
        if (start - timedelta(days=int(lead_days or 0))) <= today <= end and mult > best_mult:
            best_mult, best_name = mult, name
    return best_mult, best_name


# ─────────────────────────────────────────────────────────────────────────
# Merge + reporte
# ─────────────────────────────────────────────────────────────────────────

def merge_all(ml_results: dict, amz_results: dict) -> dict:
    agg = {}

    def add(sku, units, revenue, title, platform, account):
        if not sku or units <= 0:
            return
        row = agg.setdefault(sku, {"units": 0, "revenue": 0.0, "title": "", "platforms": set(), "accounts": set()})
        row["units"] += units
        row["revenue"] += revenue
        if title and not row["title"]:
            row["title"] = title
        row["platforms"].add(platform)
        row["accounts"].add(account)

    for nick, skus in ml_results.items():
        for sku, d in skus.items():
            add(sku, d["units"], d["revenue"], d["title"], "ml", nick)
    for nick, skus in amz_results.items():
        for sku, d in skus.items():
            add(sku, d["units"], d["revenue"], d.get("title", ""), "amazon", nick)
    return agg


def build_rows(conn, agg: dict, top_n: int, period_days: int, min_units: int = 0) -> list:
    rows = []
    for sku, d in agg.items():
        units = d["units"]
        avg_price = round(d["revenue"] / units, 2) if units else 0
        avail, reserved = bm_stock(conn, sku)
        rows.append({
            "sku": sku,
            "descripcion": d["title"] or "",
            "plataformas": "+".join(sorted(d["platforms"])),
            "cuentas": ", ".join(sorted(d["accounts"])),
            "cantidad_vendida": units,
            "precio_promedio_mxn": avg_price,
            "ingresos_mxn": round(d["revenue"], 2),
            "stock_bm_disponible": avail if avail is not None else "",
            "stock_bm_reservado": reserved if reserved is not None else "",
        })
    rows.sort(key=lambda r: r["cantidad_vendida"], reverse=True)
    if min_units > 0:
        top = [r for r in rows if r["cantidad_vendida"] >= min_units]
        if top_n:
            top = top[:top_n]
    else:
        top = rows[:top_n]
    for r in top:
        avail_n = r["stock_bm_disponible"] if isinstance(r["stock_bm_disponible"], int) else 0
        daily_rate = r["cantidad_vendida"] / period_days if period_days else 0
        boost_mult, boost_name = seasonal_boost(conn, r["descripcion"])
        demand_90d = round(daily_rate * 90 * boost_mult)
        r["demanda_proyectada_90d"] = demand_90d
        r["cantidad_sugerida_compra"] = max(demand_90d - avail_n, 0)
        if boost_mult > 1.0:
            r["descripcion"] = f"{r['descripcion']} [temporada: {boost_name}]"
    return top


def write_csv(rows: list, out_path: Path):
    fields = ["sku", "descripcion", "plataformas", "cuentas", "cantidad_vendida",
              "precio_promedio_mxn", "ingresos_mxn", "stock_bm_disponible", "stock_bm_reservado",
              "demanda_proyectada_90d", "cantidad_sugerida_compra"]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(rows: list, out_path: Path, note: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    headers_map = {
        "sku": "SKU", "descripcion": "Descripción", "plataformas": "Plataformas", "cuentas": "Cuentas",
        "cantidad_vendida": "Cant. Vendida", "precio_promedio_mxn": "Precio Prom. MXN",
        "ingresos_mxn": "Ingresos MXN", "stock_bm_disponible": "Stock BM Disp.",
        "stock_bm_reservado": "Stock BM Reserv.", "demanda_proyectada_90d": "Demanda Proy. 90d",
        "cantidad_sugerida_compra": "Cant. Sugerida Compra",
    }
    keys = list(headers_map.keys())
    int_fields = {"cantidad_vendida", "stock_bm_disponible", "stock_bm_reservado",
                  "demanda_proyectada_90d", "cantidad_sugerida_compra"}
    money_fields = {"precio_promedio_mxn", "ingresos_mxn"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de Compra"
    ws.append(list(headers_map.values()))

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c_idx in range(1, len(keys) + 1):
        c = ws.cell(row=1, column=c_idx)
        c.fill, c.font, c.border = header_fill, header_font, border
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(keys, start=1):
            val = row.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if key in money_fields:
                cell.number_format = '"$"#,##0.00'
            if key in int_fields:
                cell.alignment = Alignment(horizontal="center")
            if key == "cantidad_sugerida_compra" and isinstance(val, int) and val > 0:
                cell.font = Font(bold=True, color="B45309")

    widths = {"sku": 14, "descripcion": 45, "plataformas": 12, "cuentas": 35,
              "cantidad_vendida": 16, "precio_promedio_mxn": 16, "ingresos_mxn": 16,
              "stock_bm_disponible": 14, "stock_bm_reservado": 14,
              "demanda_proyectada_90d": 16, "cantidad_sugerida_compra": 18}
    for i, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(keys))}{len(rows) + 1}"

    note_row = len(rows) + 3
    ws.cell(row=note_row, column=1, value=note)
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="6B7280")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(keys))
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────

async def run(date_from: date, date_to: date, top_n: int, amazon_live: bool, amazon_timeout: int, min_units: int = 0):
    period_days = (date_to - date_from).days + 1
    sel = f"min {min_units} uds" if min_units > 0 else f"top {top_n}"
    print(f"=== Orden de compra: {date_from} -> {date_to} ({period_days} días), {sel} ===")

    ml_results = await fetch_ml_all_accounts(date_from, date_to)

    conn = sqlite3.connect(DB_PATH)
    amz_cache_results, amz_meta = amazon_from_cache(conn)

    amz_results = amz_cache_results
    amz_source_note = (
        f"Amazon: histórico cacheado (order_history), cobertura {amz_meta['min_date']} a "
        f"{amz_meta['max_date']}, {amz_meta['n_accounts']} cuenta(s)."
    )
    if amazon_live:
        live = await fetch_amazon_all_accounts_live(date_from, date_to, amazon_timeout)
        if live:
            # Para las cuentas que sí trajeron datos en vivo, reemplaza el cache (más preciso
            # para el rango pedido); las que fallaron quedan con lo que había en cache.
            amz_results = {**amz_cache_results, **live}
            live_nicks = ", ".join(live.keys())
            amz_source_note = (
                f"Amazon: en vivo para [{live_nicks}] en el rango {date_from}-{date_to}; "
                f"el resto (si aplica) usa cache con cobertura {amz_meta['min_date']} a {amz_meta['max_date']}."
            )

    agg = merge_all(ml_results, amz_results)
    rows = build_rows(conn, agg, top_n, period_days, min_units)

    label = f"{date_from.isoformat()}_a_{date_to.isoformat()}"
    sel_tag = f"min{min_units}" if min_units > 0 else f"top{top_n}"
    csv_path = BASE / f"orden_compra_{sel_tag}_{label}.csv"
    xlsx_path = BASE / f"orden_compra_{sel_tag}_{label}.xlsx"
    write_csv(rows, csv_path)
    note = (
        f"ML: en vivo, {date_from} a {date_to}, todas las cuentas, paginación adaptativa "
        f"(evita el límite de 10,000 resultados de la API de ML). {amz_source_note} "
        f"Demanda Proy. 90d = (ventas del periodo / {period_days}) x 90. "
        f"Cantidad sugerida = demanda proyectada 90d - stock BM disponible (piso 0)."
    )
    write_xlsx(rows, xlsx_path, note)

    print(f"\nOK -> {csv_path.name} / {xlsx_path.name} ({len(rows)} filas)")
    print(note)


def main():
    p = argparse.ArgumentParser(description="Genera orden de compra top-N SKUs más vendidos (ML+Amazon).")
    p.add_argument("--since", type=str, default=None, help="Fecha inicio YYYY-MM-DD (default: 1-ene del año actual)")
    p.add_argument("--days", type=int, default=None, help="Alternativa a --since: últimos N días")
    p.add_argument("--top", type=int, default=None, help="Cuántos SKUs incluir (default: 50 si no se usa --min-units; sin límite si se usa --min-units, salvo que se pase explícito)")
    p.add_argument("--min-units", type=int, default=0, help="En vez de top-N, incluye TODOS los SKUs con >= N unidades vendidas en el periodo")
    p.add_argument("--amazon-live", action="store_true", help="Intenta refrescar Amazon en vivo (puede tardar/fallar)")
    p.add_argument("--amazon-timeout", type=int, default=90, help="Presupuesto de segundos para Amazon en vivo (default 90)")
    args = p.parse_args()

    today = date.today()
    if args.days:
        date_from = today - timedelta(days=args.days - 1)
    elif args.since:
        date_from = datetime.strptime(args.since, "%Y-%m-%d").date()
    else:
        date_from = date(today.year, 1, 1)
    date_to = today

    top_n = args.top
    if top_n is None:
        top_n = 0 if args.min_units > 0 else 50

    asyncio.run(run(date_from, date_to, top_n, args.amazon_live, args.amazon_timeout, args.min_units))


if __name__ == "__main__":
    main()
